import streamlit as st
from datetime import datetime
from io import BytesIO
import openpyxl as _xl
from openpyxl.styles import Font as _F, PatternFill as _P, Alignment as _A
import db, calculos, exportar, importador

COLS_DB = [
    "btc","producto","ncm","proveedor","despachante","incoterm","origen","bloque","co",
    "tc","cantidad_kg","precio_unit","fob","flete","cfr","seguro","cif",
    "ajuste_incluir","ajuste_deducir","valor_aduana","di","te","base_imponible",
    "iva","iva_ad","ganancias","ing_brutos","arancel_sim","multa",
    "total_vep_usd","total_vep_ars","forwarder","terminal","acarreo","custodia",
    "senasa","senasa_madera","inal","gastos_op","honorarios","vep_anmat",
    "gastos_bancarios","lakaut","subtotal_gastos","transferencia_despa",
    "precio_total","costo_kg",
]

def _modelo_excel():
    modelo_cols = ["Referencia","Producto (NCM)","Proveedor","Cantidad (kg)",
                   "Precio Unit. (USD/kg)","Incoterm","Puerto Origen","NCM",
                   "ETD","ETA","TT Días","Estado"]
    wb = _xl.Workbook(); ws = wb.active; ws.title = "Operaciones"
    fill = _P("solid", start_color="1F3864")
    font = _F(bold=True, color="FFFFFF", name="Arial", size=10)
    for ci, h in enumerate(modelo_cols, 1):
        c = ws.cell(1, ci, h); c.fill = fill; c.font = font
        c.alignment = _A(horizontal="center")
    ws.append(["BTC-2590","ALGINATO DE SODIO","Qingdao Bright Moon",5000,
                10.10,"FOB","Qingdao","3913.10.00.210P","28/03/2026","14/05/2026",47,"Confirmado"])
    buf = BytesIO(); wb.save(buf)
    return buf.getvalue()

def render():
    st.caption("Subi el Excel de seguimiento mensual y la app calcula todos los costeos de una.")

    cfg = db.get_all_config()
    tc = float(cfg.get("tc", 1400))
    distancia_km = int(cfg.get("distancia_km", 50))

    st.download_button("📄 Descargar modelo Excel", _modelo_excel(),
        file_name="modelo_importacion.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    archivo = st.file_uploader("Excel de operaciones (.xlsx)", type=["xlsx"])
    if not archivo:
        return

    filas = importador.normalizar_filas(importador.leer_excel(archivo.read()))
    if not filas:
        st.error("No se encontraron filas válidas.")
        return

    st.success(f"Se encontraron **{len(filas)}** operaciones. Revisá antes de importar.")
    st.divider()

    filas_edit = []
    for i, fila in enumerate(filas):
        with st.expander(f"**{fila['btc']}** — {fila['producto']} | {fila['proveedor']} | {fila['incoterm']} {fila['origen']}", expanded=False):
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                co = st.selectbox("C.O.", ["NO","SI"], index=1 if fila["co"]=="SI" else 0, key=f"co_{i}")
            with col2:
                despa = st.selectbox("Despachante", ["Mestre","Adduci"],
                                     index=0 if fila["despachante"]=="Mestre" else 1, key=f"despa_{i}")
            with col3:
                bloque = st.selectbox("Bloque", ["NO MERCOSUR","MERCOSUR","ALADI"],
                                      index=["NO MERCOSUR","MERCOSUR","ALADI"].index(fila["bloque"]),
                                      key=f"bloque_{i}")
            with col4:
                dist = st.selectbox("Distancia", [0,30,50,70,100],
                                    index=[0,30,50,70,100].index(distancia_km) if distancia_km in [0,30,50,70,100] else 2,
                                    format_func=lambda x:{0:"CABA",30:"GBA <=30km",50:"GBA <=50km",70:"GBA <=70km",100:"Pilar"}[x],
                                    key=f"dist_{i}")
            col1, col2 = st.columns(2)
            with col1:
                ncm_display = fila["ncm"] or "se tomara de posiciones"
                st.caption(f"NCM: {ncm_display}")
                st.caption(f"KG: {fila['cantidad_kg']:,.0f} | Precio: USD {fila['precio_unit']:,.2f} | FOB: USD {fila['cantidad_kg']*fila['precio_unit']:,.2f}")
            with col2:
                st.caption(f"ETD: {fila['etd']} | ETA: {fila['eta']} | TT: {fila['tt_dias']} dias")
            filas_edit.append({**fila, "co": co, "despachante": despa, "bloque": bloque, "distancia_km": dist})

    st.divider()
    importar_btn = st.button("💾 Importar todos", type="primary")

    if importar_btn:
        posiciones_dict = {p["producto"]: p for p in db.get_posiciones()}
        errores = []; guardados = 0
        progress = st.progress(0)

        for idx, fila in enumerate(filas_edit):
            try:
                pos = posiciones_dict.get(fila["producto"], {})
                ncm = fila["ncm"] or pos.get("ncm", "")
                if not ncm:
                    for pk, pd_ in posiciones_dict.items():
                        if fila["producto"].upper() in pk.upper() or pk.upper() in fila["producto"].upper():
                            pos = pd_; ncm = pd_.get("ncm", ""); break

                resultado = calculos.calcular_costeo({
                    "btc": fila["btc"], "producto": fila["producto"], "ncm": ncm,
                    "proveedor": fila["proveedor"], "despachante": fila["despachante"],
                    "incoterm": fila["incoterm"] or "FOB", "origen": fila["origen"] or "Otro",
                    "bloque": fila["bloque"], "co": fila["co"], "tc": tc,
                    "cantidad_kg": fila["cantidad_kg"], "precio_unit": fila["precio_unit"],
                    "flete": fila["flete"], "ajuste_incluir": fila["ajuste_incluir"],
                    "ajuste_deducir": fila["ajuste_deducir"],
                    "die_pct": pos.get("die",0), "te_pct": pos.get("te",0.03),
                    "iva_pct": pos.get("iva",0.21), "iva_ad_ncm": pos.get("iva_ad","NO"),
                    "iva_ad_pct": pos.get("iva_ad_pct",0.20),
                    "ganancias_ncm": pos.get("ganancias","NO"), "ganancias_pct": pos.get("ganancias_pct",0.06),
                    "distancia_km": fila["distancia_km"], "multa": fila["multa"],
                })
                data = {k: resultado[k] for k in COLS_DB}
                data["fecha"] = fila.get("etd") or datetime.now().strftime("%Y-%m-%d %H:%M")
                data["notas"] = f"Importado lote | Estado: {fila['estado']} | ETD: {fila['etd']} | ETA: {fila['eta']}"
                db.save_costeo(data)
                guardados += 1
            except Exception as e:
                errores.append(f"{fila['btc']}: {e}")
            progress.progress((idx + 1) / len(filas_edit))

        st.success(f"✅ {guardados} costeos importados.")
        if errores:
            st.warning(f"⚠️ {len(errores)} con advertencias:")
            for e in errores:
                st.caption(e)

        if guardados > 0:
            st.divider()
            st.markdown("**Descargar Excel por costeo:**")
            for fila in filas_edit:
                c = db.get_costeo_by_btc(fila["btc"])
                if c:
                    xlsx = exportar.generar_excel(c, guardar_local=True)
                    st.download_button(f"📊 {fila['btc']} — {fila['producto']}", data=xlsx,
                        file_name=f"SDF_{fila['btc']}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_{fila['btc']}")
