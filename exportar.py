"""
exportar.py - Genera reporte Excel + PDF de un costeo BIOTEC SA
Resumen: números puros con fórmulas editables
Calculadora: hoja de inputs separada
"""
from pathlib import Path
from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

REPORTES_DIR = Path(__file__).parent / "reportes"
REPORTES_DIR.mkdir(exist_ok=True)

# ── Colores ───────────────────────────────────────────────────────
AZUL_OSCURO  = "1F3864"
AZUL_MED     = "2E75B6"
VERDE_OSCURO = "375623"
VERDE_CLARO  = "E2EFDA"
GRIS_CLARO   = "F2F2F2"
GRIS_MED     = "D9D9D9"
NARANJA      = "E65100"
BLANCO       = "FFFFFF"
AMARILLO     = "FFF9C4"  # inputs editables

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Arial")

def _border_thin():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _usd(val):
    try: return float(val)
    except: return 0.0

def _set_num(cell, value, fmt="#,##0.00", bold=False, color="000000", bg=None):
    cell.value = value
    cell.number_format = fmt
    cell.font = _font(bold=bold, color=color)
    cell.alignment = _align("right")
    cell.border = _border_thin()
    if bg:
        cell.fill = _fill(bg)

def _set_label(cell, text, bold=True, color="444444", bg=GRIS_CLARO):
    cell.value = text
    cell.font = _font(bold=bold, color=color, size=9)
    cell.fill = _fill(bg)
    cell.border = _border_thin()
    cell.alignment = _align("left")

def _seccion(ws, row, titulo, color):
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row, 1, titulo)
    c.fill = _fill(color)
    c.font = _font(True, BLANCO, 10)
    c.alignment = _align("center")
    ws.row_dimensions[row].height = 16

def _fila(ws, row, lbl1, val1, lbl2, val2, fmt="#,##0.00", bg_val=BLANCO, input_cell=False):
    """Fila con dos pares label/valor. Si input_cell=True, fondo amarillo."""
    bg = AMARILLO if input_cell else bg_val
    _set_label(ws.cell(row, 1), lbl1)
    _set_num(ws.cell(row, 2), val1, fmt=fmt, bg=bg)
    _set_label(ws.cell(row, 3), lbl2)
    _set_num(ws.cell(row, 4), val2, fmt=fmt, bg=bg)
    ws.row_dimensions[row].height = 15


# ════════════════════════════════════════════════════════════════════
#  EXCEL
# ════════════════════════════════════════════════════════════════════

def generar_excel(costeo: dict, guardar_local=True) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _hoja_resumen(wb, costeo)
    _hoja_calculadora(wb, costeo)
    wb.calculation.calcMode = "auto"

    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    if guardar_local:
        btc = costeo.get("btc", "SIN_BTC").replace("/", "-")
        fecha = datetime.now().strftime("%Y%m%d_%H%M")
        path = REPORTES_DIR / f"SDF_{btc}_{fecha}.xlsx"
        path.write_bytes(data)

    return data


def _hoja_resumen(wb, c):
    ws = wb.create_sheet("Resumen")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16

    row = 1
    # Header
    ws.merge_cells(f"A{row}:D{row}")
    h = ws.cell(row, 1, "BIOTEC S.A. - COSTEO DE IMPORTACIÓN")
    h.fill = _fill(AZUL_OSCURO); h.font = _font(True, BLANCO, 14)
    h.alignment = _align("center"); ws.row_dimensions[row].height = 28
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ts = ws.cell(row, 1, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    ts.fill = _fill(AZUL_MED); ts.font = _font(False, BLANCO, 9)
    ts.alignment = _align("right"); row += 2

    # A · Datos generales (texto)
    _seccion(ws, row, "A · DATOS GENERALES", AZUL_MED); row += 1
    datos = [
        ("BTC", c.get("btc",""), "Producto", c.get("producto","")),
        ("Proveedor", c.get("proveedor",""), "NCM", c.get("ncm","")),
        ("Despachante", c.get("despachante",""), "Incoterm", c.get("incoterm","")),
        ("Origen", c.get("origen",""), "Bloque", c.get("bloque","")),
        ("C.O.", c.get("co",""), "Notas", c.get("notas","")),
        ("ETD", c.get("etd",""), "ETA", c.get("eta","")),
    ]
    for l1, v1, l2, v2 in datos:
        for col, val in [(1,l1),(2,v1),(3,l2),(4,v2)]:
            cell = ws.cell(row, col, val)
            cell.font = _font(bold=(col in [1,3]), color="444444", size=9)
            cell.fill = _fill(GRIS_CLARO if col in [1,3] else BLANCO)
            cell.border = _border_thin()
        ws.row_dimensions[row].height = 15; row += 1
    row += 1

    # B · Inputs editables (amarillo) — KG, precio, flete, TC
    _seccion(ws, row, "B · INPUTS EDITABLES  ★ modificá estas celdas", AZUL_MED); row += 1

    # Guardar referencias de inputs para fórmulas
    r_kg    = row; _fila(ws, row, "Cantidad KG", _usd(c.get("cantidad_kg")), "TC USD→ARS", _usd(c.get("tc",1400)), input_cell=True); row += 1
    r_pu    = row; _fila(ws, row, "Precio unit (USD)", _usd(c.get("precio_unit")), "Flete total (USD)", _usd(c.get("flete")), input_cell=True); row += 1
    r_ajinc = row; _fila(ws, row, "Ajuste incluir (USD)", _usd(c.get("ajuste_incluir")), "Ajuste deducir (USD)", _usd(c.get("ajuste_deducir")), input_cell=True); row += 1

    # Porcentajes
    die_pct = _usd(c.get("di",0)) / (_usd(c.get("valor_aduana",1)) or 1)
    te_pct  = _usd(c.get("te",0))  / (_usd(c.get("valor_aduana",1)) or 1)
    iva_pct = _usd(c.get("iva",0)) / (_usd(c.get("base_imponible",1)) or 1)
    iva_ad_pct = _usd(c.get("iva_ad",0)) / (_usd(c.get("base_imponible",1)) or 1)
    gan_pct = _usd(c.get("ganancias",0)) / (_usd(c.get("base_imponible",1)) or 1)

    r_die   = row; _fila(ws, row, "DIE %", die_pct, "TE %", te_pct, fmt="0.00%", input_cell=True); row += 1
    r_iva   = row; _fila(ws, row, "IVA %", iva_pct, "IVA AD %", iva_ad_pct, fmt="0.00%", input_cell=True); row += 1
    r_gan   = row; _fila(ws, row, "Ganancias %", gan_pct, "Arancel SIM (USD)", _usd(c.get("arancel_sim",10)), input_cell=True); row += 1
    r_multa = row; _fila(ws, row, "Multa (USD)", _usd(c.get("multa")), "", "", input_cell=True); row += 1
    row += 1

    # C · Valor mercadería (fórmulas)
    _seccion(ws, row, "C · VALOR MERCADERÍA (calculado)", AZUL_MED); row += 1

    kg_ref  = f"B{r_kg}"; pu_ref = f"D{r_kg}"; flete_ref = f"D{r_pu}"
    tc_ref  = f"D{r_kg}"; ajinc_ref = f"B{r_ajinc}"; ajded_ref = f"D{r_ajinc}"
    die_ref = f"B{r_die}"; te_ref_pct = f"D{r_die}"
    iva_ref_pct = f"B{r_iva}"; ivaad_ref_pct = f"D{r_iva}"
    gan_ref_pct = f"B{r_gan}"; arancel_ref = f"D{r_gan}"
    multa_ref = f"B{r_multa}"

    fob_row = row
    _set_label(ws.cell(row, 1), "FOB (USD)")
    ws.cell(row, 2).value = f"={kg_ref}*B{r_pu}"
    ws.cell(row, 2).number_format = "#,##0.00"; ws.cell(row, 2).border = _border_thin(); ws.cell(row, 2).alignment = _align("right")
    _set_label(ws.cell(row, 3), "Flete (USD)")
    ws.cell(row, 4).value = flete_ref  # referencia directa
    ws.cell(row, 4).value = f"={flete_ref}"
    ws.cell(row, 4).number_format = "#,##0.00"; ws.cell(row, 4).border = _border_thin(); ws.cell(row, 4).alignment = _align("right")
    ws.row_dimensions[row].height = 15; row += 1

    seg_row = row
    _set_label(ws.cell(row, 1), "Seguro (USD)")
    ws.cell(row, 2).value = f"=(B{fob_row}+D{fob_row})*0.005"
    ws.cell(row, 2).number_format = "#,##0.00"; ws.cell(row, 2).border = _border_thin(); ws.cell(row, 2).alignment = _align("right")
    _set_label(ws.cell(row, 3), "CIF (USD)")
    ws.cell(row, 4).value = f"=B{fob_row}+D{fob_row}+B{seg_row}"
    ws.cell(row, 4).number_format = "#,##0.00"; ws.cell(row, 4).border = _border_thin(); ws.cell(row, 4).alignment = _align("right")
    ws.row_dimensions[row].height = 15; row += 1

    va_row = row
    _set_label(ws.cell(row, 1), "Valor Aduana (USD)")
    ws.cell(row, 2).value = f"=D{seg_row}+{ajinc_ref}-{ajded_ref}"
    ws.cell(row, 2).number_format = "#,##0.00"; ws.cell(row, 2).border = _border_thin(); ws.cell(row, 2).alignment = _align("right"); ws.cell(row, 2).font = _font(bold=True, color=AZUL_MED)
    _set_label(ws.cell(row, 3), "Base Imponible (USD)")
    ws.cell(row, 4).value = f"=B{va_row}*(1+{die_ref}+{te_ref_pct})"
    ws.cell(row, 4).number_format = "#,##0.00"; ws.cell(row, 4).border = _border_thin(); ws.cell(row, 4).alignment = _align("right")
    ws.row_dimensions[row].height = 15; row += 1
    row += 1

    # D · VEP Aduana (fórmulas)
    _seccion(ws, row, "D · VEP ADUANA (calculado)", VERDE_OSCURO); row += 1

    tributos = [
        ("D. Importación (USD)", f"=B{va_row}*{die_ref}",         "IVA (USD)",          f"=D{va_row}*{iva_ref_pct}"),
        ("T. Estadística (USD)", f"=B{va_row}*{te_ref_pct}",      "IVA Adicional (USD)", f"=D{va_row}*{ivaad_ref_pct}"),
        ("Ing. Brutos (USD)",    f"=D{va_row}*0.025",              "Imp. Ganancias (USD)",f"=D{va_row}*{gan_ref_pct}"),
        ("Arancel SIM (USD)",    f"={arancel_ref}",                "Multa (USD)",         f"={multa_ref}"),
    ]
    trib_rows = []
    for l1, f1, l2, f2 in tributos:
        trib_rows.append(row)
        _set_label(ws.cell(row, 1), l1, bg=VERDE_CLARO)
        for col, formula in [(2, f1), (4, f2)]:
            ws.cell(row, col).value = formula
            ws.cell(row, col).number_format = "#,##0.00"
            ws.cell(row, col).border = _border_thin()
            ws.cell(row, col).fill = _fill(VERDE_CLARO)
            ws.cell(row, col).alignment = _align("right")
        _set_label(ws.cell(row, 3), l2, bg=VERDE_CLARO)
        ws.row_dimensions[row].height = 15; row += 1

    # Total VEP
    vep_formula = f"=B{trib_rows[0]}+B{trib_rows[1]}+B{trib_rows[2]}+B{trib_rows[3]}+D{trib_rows[0]}+D{trib_rows[1]}+D{trib_rows[2]}+D{trib_rows[3]}"
    ws.merge_cells(f"A{row}:B{row}")
    lv = ws.cell(row, 1, "TOTAL VEP (USD)")
    lv.fill = _fill(VERDE_OSCURO); lv.font = _font(True, BLANCO, 11); lv.alignment = _align("center")
    vep_row = row
    ws.cell(row, 3).value = vep_formula
    ws.cell(row, 3).number_format = "#,##0.00"; ws.cell(row, 3).font = _font(True, VERDE_OSCURO, 11); ws.cell(row, 3).alignment = _align("right")
    ws.cell(row, 4).value = f"=C{vep_row}*{tc_ref}"
    ws.cell(row, 4).number_format = "#,##0.00"; ws.cell(row, 4).font = _font(True, VERDE_OSCURO, 11); ws.cell(row, 4).alignment = _align("right")
    ws.row_dimensions[row].height = 20; row += 2

    # E · Gastos locales (inputs editables)
    _seccion(ws, row, "E · GASTOS LOCALES  ★ editables", AZUL_MED); row += 1

    gastos = [
        ("", 0,                                               "Terminal",          _usd(c.get("terminal"))),
        ("Acarreo",          _usd(c.get("acarreo")),         "Custodia",          _usd(c.get("custodia"))),
        ("SENASA",           _usd(c.get("senasa")),           "SENASA Madera",     _usd(c.get("senasa_madera"))),
        ("Gastos Operativos",_usd(c.get("gastos_op")),       "G. Bancarios",      _usd(c.get("gastos_bancarios"))),
        ("Honorarios",       _usd(c.get("honorarios")),      "Lakaut",            _usd(c.get("lakaut"))),
        ("INAL ← BIOTEC",    _usd(c.get("inal")),             "VEP ANMAT ← BIOTEC",_usd(c.get("vep_anmat"))),
    ]
    gasto_rows = []
    for l1, v1, l2, v2 in gastos:
        gasto_rows.append(row)
        is_biotec = "BIOTEC" in l1 or "BIOTEC" in l2
        _set_label(ws.cell(row, 1), l1, color=NARANJA if is_biotec else "444444")
        _set_num(ws.cell(row, 2), v1, bg=AMARILLO)
        _set_label(ws.cell(row, 3), l2, color=NARANJA if is_biotec else "444444")
        _set_num(ws.cell(row, 4), v2, bg=AMARILLO)
        ws.row_dimensions[row].height = 15; row += 1

    # Subtotal gastos
    sum_parts = "+".join([f"B{r}+D{r}" for r in gasto_rows])
    subtotal_row = row
    ws.merge_cells(f"A{row}:B{row}")
    ls = ws.cell(row, 1, "SUBTOTAL GASTOS (USD)")
    ls.fill = _fill(GRIS_MED); ls.font = _font(True, size=10); ls.alignment = _align("center")
    ws.cell(row, 3).value = f"={sum_parts}"
    ws.cell(row, 3).number_format = "#,##0.00"; ws.cell(row, 3).font = _font(True, size=10); ws.cell(row, 3).alignment = _align("right")
    ws.row_dimensions[row].height = 18; row += 1

    # Adelanto al despachante (suma VEP + gastos sin BIOTEC)
    adelanto_row = row
    ws.merge_cells(f"A{row}:B{row}")
    la = ws.cell(row, 1, "ADELANTO AL DESPACHANTE (USD)")
    la.fill = _fill(AZUL_MED); la.font = _font(True, BLANCO, 10); la.alignment = _align("center")
    # VEP + gastos excepto INAL y VEP ANMAT (fila 5 = gasto_rows[-1])
    gastos_despa = "+".join([f"B{r}+D{r}" for r in gasto_rows[:-1]])  # excluye última fila (INAL/ANMAT)
    ws.cell(row, 3).value = f"=C{vep_row}+{gastos_despa}"
    ws.cell(row, 3).number_format = "#,##0.00"; ws.cell(row, 3).font = _font(True, BLANCO, 10)
    ws.cell(row, 3).fill = _fill(AZUL_MED); ws.cell(row, 3).alignment = _align("right")
    ws.cell(row, 4).value = f"=C{adelanto_row}*{tc_ref}"
    ws.cell(row, 4).number_format = "#,##0.00"; ws.cell(row, 4).font = _font(True, BLANCO, 10)
    ws.cell(row, 4).fill = _fill(AZUL_MED); ws.cell(row, 4).alignment = _align("right")
    ws.row_dimensions[row].height = 18; row += 2

    # F · Resumen final
    _seccion(ws, row, "F · RESUMEN FINAL", AZUL_OSCURO); row += 1
    totales = [
        ("Precio Total (USD)", f"=B{va_row}+C{vep_row}+C{subtotal_row}", "Precio Total (ARS)", None),
        ("Costo x KG (USD)",   f"=IF({kg_ref}>0,B{row-1}/{kg_ref},0)",    "Costo x KG (ARS)",   None),
    ]
    total_rows = []
    for l1, f1, l2, _ in totales:
        total_rows.append(row)
        _set_label(ws.cell(row, 1), l1, bg=GRIS_CLARO)
        ws.cell(row, 2).value = f1
        ws.cell(row, 2).number_format = "#,##0.00"; ws.cell(row, 2).font = _font(True, color=AZUL_MED)
        ws.cell(row, 2).border = _border_thin(); ws.cell(row, 2).alignment = _align("right")
        _set_label(ws.cell(row, 3), l2, bg=GRIS_CLARO)
        ws.cell(row, 4).value = f"=B{row}*{tc_ref}"
        ws.cell(row, 4).number_format = "#,##0.00"; ws.cell(row, 4).font = _font(True, color=AZUL_MED)
        ws.cell(row, 4).border = _border_thin(); ws.cell(row, 4).alignment = _align("right")
        ws.row_dimensions[row].height = 18; row += 1

    # Pie
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    pie = ws.cell(row, 1, f"★ Celdas en amarillo son editables | Notas: {c.get('notas','')} | Estado: {c.get('estado','')}")
    pie.font = _font(italic=True, size=8, color="888888")


def _hoja_calculadora(wb, c):
    """Hoja Calculadora — inputs en col B, fórmulas en col D (igual que antes)."""
    ws = wb.create_sheet("Calculadora")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 18

    ws.merge_cells("A1:D1")
    h = ws.cell(1, 1, f"CALCULADORA - {c.get('btc','')} - {c.get('producto','')}")
    h.fill = _fill(AZUL_OSCURO); h.font = _font(True, BLANCO, 12)
    h.alignment = _align("center"); ws.row_dimensions[1].height = 24

    row = 2
    _seccion(ws, row, "INPUTS — editar estas celdas para recalcular", AZUL_MED); row += 1

    die_pct = _usd(c.get("di",0)) / (_usd(c.get("valor_aduana",1)) or 1)
    te_pct  = _usd(c.get("te",0))  / (_usd(c.get("valor_aduana",1)) or 1)
    iva_pct = _usd(c.get("iva",0)) / (_usd(c.get("base_imponible",1)) or 1)
    iva_ad_pct = _usd(c.get("iva_ad",0)) / (_usd(c.get("base_imponible",1)) or 1)
    gan_pct = _usd(c.get("ganancias",0)) / (_usd(c.get("base_imponible",1)) or 1)

    inputs = [
        ("KG",               c.get("cantidad_kg", 0),    "kg",         "#,##0.00"),
        ("Precio unit (USD)", c.get("precio_unit", 0),   "precio_unit","#,##0.00"),
        ("Flete (USD)",       c.get("flete", 0),          "flete",      "#,##0.00"),
        ("TC USD→ARS",        c.get("tc", 1400),          "tc",         "#,##0.00"),
        ("Ajuste incluir",    c.get("ajuste_incluir", 0), "aj_inc",     "#,##0.00"),
        ("Ajuste deducir",    c.get("ajuste_deducir", 0), "aj_ded",     "#,##0.00"),
        ("DIE %",             die_pct,                    "die_pct",    "0.00%"),
        ("TE %",              te_pct,                     "te_pct",     "0.00%"),
        ("IVA %",             iva_pct,                    "iva_pct",    "0.00%"),
        ("IVA AD %",          iva_ad_pct,                 "iva_ad_pct", "0.00%"),
        ("Ganancias %",       gan_pct,                    "gan_pct",    "0.00%"),
        ("Arancel SIM",       c.get("arancel_sim", 10),   "arancel_sim","#,##0.00"),
        ("Multa",             c.get("multa", 0),           "multa",      "#,##0.00"),
        ("Terminal",          c.get("terminal", 0),        "terminal",   "#,##0.00"),
        ("Acarreo",           c.get("acarreo", 0),         "acarreo",    "#,##0.00"),
        ("SENASA",            c.get("senasa", 0),           "senasa",     "#,##0.00"),
        ("SENASA Madera",     c.get("senasa_madera", 0),   "senasa_mad", "#,##0.00"),
        ("INAL",              c.get("inal", 0),             "inal",       "#,##0.00"),
        ("Gastos Operativos", c.get("gastos_op", 0),       "gastos_op",  "#,##0.00"),
        ("Honorarios",        c.get("honorarios", 0),      "honorarios", "#,##0.00"),
        ("VEP ANMAT",         c.get("vep_anmat", 0),       "vep_anmat",  "#,##0.00"),
        ("G. Bancarios",      c.get("gastos_bancarios", 0),"g_ban",      "#,##0.00"),
        ("Lakaut",            c.get("lakaut", 0),           "lakaut",     "#,##0.00"),
    ]

    ref = {}
    for label, value, key, fmt in inputs:
        lc = ws.cell(row, 1, label)
        lc.fill = _fill(AMARILLO); lc.font = _font(True, "444444", 9); lc.border = _border_thin()
        vc = ws.cell(row, 2, value)
        vc.number_format = fmt; vc.font = _font(False, "0000FF", 9)
        vc.border = _border_thin(); vc.alignment = _align("right")
        ref[key] = f"B{row}"; row += 1

    row += 1
    _seccion(ws, row, "CÁLCULOS — generados por fórmulas", VERDE_OSCURO); row += 1

    def calc_row(label, formula, highlight=False):
        nonlocal row
        lc = ws.cell(row, 3, label)
        lc.fill = _fill(VERDE_CLARO if not highlight else AMARILLO)
        lc.font = _font(True, "444444", 9); lc.border = _border_thin()
        vc = ws.cell(row, 4)
        vc.value = formula; vc.number_format = "#,##0.00"
        vc.font = _font(highlight, "000000" if not highlight else VERDE_OSCURO, 9)
        vc.border = _border_thin(); vc.alignment = _align("right")
        result_row = row; row += 1
        return f"D{result_row}"

    r = ref
    fob_ref    = calc_row("FOB (USD)",             f"={r['kg']}*{r['precio_unit']}")
    seg_ref    = calc_row("Seguro (USD)",           f"=({fob_ref}+{r['flete']})*0.005")
    cif_ref    = calc_row("CIF (USD)",              f"={fob_ref}+{r['flete']}+{seg_ref}")
    va_ref     = calc_row("Valor Aduana (USD)",     f"={cif_ref}+{r['aj_inc']}-{r['aj_ded']}")
    di_ref     = calc_row("D. Importación (USD)",   f"={va_ref}*{r['die_pct']}")
    te_ref     = calc_row("T. Estadística (USD)",   f"={va_ref}*{r['te_pct']}")
    base_ref   = calc_row("Base Imponible (USD)",   f"={va_ref}+{di_ref}+{te_ref}")
    iva_ref    = calc_row("IVA (USD)",              f"={base_ref}*{r['iva_pct']}")
    iva_ad_ref = calc_row("IVA Adicional (USD)",    f"={base_ref}*{r['iva_ad_pct']}")
    gan_ref    = calc_row("Ganancias (USD)",        f"={base_ref}*{r['gan_pct']}")
    ib_ref     = calc_row("Ing. Brutos (USD)",      f"={base_ref}*0.025")
    vep_ref    = calc_row("TOTAL VEP (USD)",        f"={di_ref}+{te_ref}+{iva_ref}+{iva_ad_ref}+{gan_ref}+{ib_ref}+{r['arancel_sim']}+{r['multa']}", highlight=True)
    calc_row("TOTAL VEP (ARS)",     f"={vep_ref}*{r['tc']}", highlight=True)
    gastos_sum = "+".join([r['terminal'],r['acarreo'],r['senasa'],r['senasa_mad'],r['inal'],r['gastos_op'],r['honorarios'],r['vep_anmat'],r['g_ban'],r['lakaut']])
    sub_ref    = calc_row("Subtotal Gastos (USD)",  f"=SUM({gastos_sum})", highlight=True)
    tot_ref    = calc_row("PRECIO TOTAL (USD)",     f"={va_ref}+{vep_ref}+{sub_ref}", highlight=True)
    calc_row("COSTO x KG (USD)",    f"=IF({r['kg']}>0,{tot_ref}/{r['kg']},0)", highlight=True)
    calc_row("PRECIO TOTAL (ARS)",  f"={tot_ref}*{r['tc']}", highlight=True)
    calc_row("COSTO x KG (ARS)",    f"=IF({r['kg']}>0,{tot_ref}/{r['kg']}*{r['tc']},0)", highlight=True)


# ════════════════════════════════════════════════════════════════════
#  PDF (sin cambios)
# ════════════════════════════════════════════════════════════════════

class PDF(FPDF):
    def header(self):
        self.set_fill_color(31, 56, 100)
        self.rect(0, 0, 210, 18, 'F')
        self.set_font("Helvetica", "B", 13)
        self.set_text_color(255, 255, 255)
        self.set_xy(10, 4)
        self.cell(0, 10, "BIOTEC S.A. - COSTEO DE IMPORTACION", align="L")
        self.set_font("Helvetica", "", 8)
        self.set_xy(10, 11)
        self.cell(0, 5, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", align="R")
        self.ln(14)

    def footer(self):
        self.set_y(-12)
        self.set_font("Helvetica", "I", 7)
        self.set_text_color(150, 150, 150)
        self.cell(0, 5, f"Pág. {self.page_no()} - Uso interno BIOTEC SA", align="C")


def generar_pdf(costeo: dict, guardar_local=True) -> bytes:
    c = costeo
    pdf = PDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_margins(12, 22, 12)

    def seccion(titulo, r, g, b):
        pdf.set_fill_color(r, g, b)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(0, 7, f"  {titulo}", fill=True, ln=True)
        pdf.set_text_color(0, 0, 0)

    def fila(lbl, val, lbl2="", val2="", highlight=False):
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_fill_color(242, 242, 242)
        if highlight:
            pdf.set_fill_color(230, 81, 0)
            pdf.set_text_color(255, 255, 255)
        pdf.cell(47, 6, f" {lbl}", fill=True)
        pdf.set_font("Helvetica", "", 8)
        pdf.set_fill_color(255, 255, 255) if not highlight else pdf.set_fill_color(255, 230, 220)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(48, 6, f" {val}", fill=True)
        if lbl2:
            pdf.set_font("Helvetica", "B", 8)
            pdf.set_fill_color(242, 242, 242)
            pdf.cell(47, 6, f" {lbl2}", fill=True)
            pdf.set_font("Helvetica", "", 8)
            pdf.set_fill_color(255, 255, 255)
            pdf.cell(0, 6, f" {val2}", fill=True)
        pdf.ln()

    def usd(val):
        try: return f"USD {float(val):,.2f}"
        except: return "-"

    def ars(val):
        try: return f"$ {float(val):,.0f}"
        except: return "-"

    def total_row(lbl, val_usd, val_ars="", r=55, g=86, b=35):
        pdf.set_fill_color(r, g, b)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(95, 7, f"  {lbl}", fill=True)
        pdf.cell(0, 7, f"  {val_usd}   {val_ars}", fill=True)
        pdf.set_text_color(0, 0, 0)
        pdf.ln()

    pdf.ln(2)
    seccion("A · DATOS GENERALES", 46, 117, 182)
    fila("BTC", c.get("btc",""), "Producto", c.get("producto",""))
    fila("Proveedor", c.get("proveedor",""), "NCM", c.get("ncm",""))
    fila("Despachante", c.get("despachante",""), "Incoterm", c.get("incoterm",""))
    fila("Origen", c.get("origen",""), "Bloque", c.get("bloque",""))
    fila("C.O.", c.get("co",""), "TC USD>ARS", f"$ {float(c.get('tc',0)):,.0f}")
    fila("Cantidad KG", f"{float(c.get('cantidad_kg',0)):,.0f}", "Precio unit", usd(c.get("precio_unit")))
    pdf.ln(2)

    seccion("B · VALOR MERCADERÍA", 46, 117, 182)
    fila("FOB", usd(c.get("fob")), "CIF", usd(c.get("cif")))
    fila("Flete", usd(c.get("flete")), "Seguro", usd(c.get("seguro")))
    fila("Valor en Aduana", usd(c.get("valor_aduana")), "Base Imponible", usd(c.get("base_imponible")))
    pdf.ln(2)

    seccion("C · VEP ADUANA", 55, 86, 35)
    fila("D. Importación", usd(c.get("di")), "IVA", usd(c.get("iva")))
    fila("T. Estadística", usd(c.get("te")), "IVA Adicional", usd(c.get("iva_ad")))
    fila("Ing. Brutos", usd(c.get("ing_brutos")), "Imp. Ganancias", usd(c.get("ganancias")))
    fila("Arancel SIM", usd(c.get("arancel_sim")), "Multa", usd(c.get("multa")))
    total_row("TOTAL VEP", usd(c.get("total_vep_usd")), ars(c.get("total_vep_ars")))
    pdf.ln(2)

    seccion("D · GASTOS LOCALES", 46, 117, 182)
    fila("Terminal/Almacen", usd(c.get("terminal")), "", "")
    fila("Acarreo", usd(c.get("acarreo")), "Custodia", usd(c.get("custodia")))
    fila("SENASA", usd(c.get("senasa")), "SENASA Madera", usd(c.get("senasa_madera")))
    fila("Gastos Operativos", usd(c.get("gastos_op")), "G. Bancarios", usd(c.get("gastos_bancarios")))
    fila("Honorarios", usd(c.get("honorarios")), "Lakaut", usd(c.get("lakaut")))
    fila("INAL <- BIOTEC", usd(c.get("inal")), "VEP ANMAT <- BIOTEC", usd(c.get("vep_anmat")), highlight=True)
    total_row("TRANSFERENCIA AL DESPACHANTE", usd(c.get("transferencia_despa")), "", 46, 117, 182)
    total_row("SUBTOTAL GASTOS", usd(c.get("subtotal_gastos")), "", 80, 80, 80)
    pdf.ln(2)

    seccion("E · RESUMEN FINAL", 31, 56, 100)
    total_row("PRECIO TOTAL", usd(c.get("precio_total")),
              ars(float(c.get("precio_total",0)) * float(c.get("tc",1))), 31, 56, 100)
    total_row("COSTO x KG", usd(c.get("costo_kg")),
              ars(float(c.get("costo_kg",0)) * float(c.get("tc",1))), 31, 56, 100)

    if c.get("notas"):
        pdf.ln(3)
        pdf.set_font("Helvetica", "I", 8)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 5, f"Notas: {c.get('notas')}", ln=True)

    buf = BytesIO()
    pdf.output(buf)
    data = buf.getvalue()

    if guardar_local:
        btc = c.get("btc", "SIN_BTC").replace("/", "-")
        fecha = datetime.now().strftime("%Y%m%d_%H%M")
        path = REPORTES_DIR / f"SDF_{btc}_{fecha}.pdf"
        path.write_bytes(data)

    return data
